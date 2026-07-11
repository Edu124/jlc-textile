import csv
import io
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy.orm import Session
from ..db import get_db
from ..auth import require_user
from .. import models, services

router = APIRouter(prefix="/api", tags=["masters"], dependencies=[Depends(require_user)])


# ══ Units ═════════════════════════════════════════════════════════════════════
class UnitIn(BaseModel):
    name: str
    abbreviation: str


@router.get("/units")
def list_units(db: Session = Depends(get_db)):
    return [{"id": u.id, "name": u.name, "abbreviation": u.abbreviation}
            for u in db.query(models.Unit).order_by(models.Unit.name).all()]


@router.post("/units")
def create_unit(body: UnitIn, db: Session = Depends(get_db)):
    u = models.Unit(name=body.name.strip(), abbreviation=body.abbreviation.strip())
    db.add(u); db.commit(); db.refresh(u)
    return {"id": u.id}


@router.put("/units/{uid}")
def update_unit(uid: int, body: UnitIn, db: Session = Depends(get_db)):
    u = db.query(models.Unit).get(uid)
    if not u: raise HTTPException(404, "Not found")
    u.name, u.abbreviation = body.name.strip(), body.abbreviation.strip()
    db.commit(); return {"ok": True}


@router.delete("/units/{uid}")
def delete_unit(uid: int, db: Session = Depends(get_db)):
    u = db.query(models.Unit).get(uid)
    if u:
        try:
            db.delete(u); db.commit()
        except Exception:
            db.rollback(); raise HTTPException(409, "Unit is in use")
    return {"ok": True}


# ══ Product categories ════════════════════════════════════════════════════════
class CategoryIn(BaseModel):
    name: str


@router.get("/categories")
def list_categories(db: Session = Depends(get_db)):
    return [{"id": c.id, "name": c.name}
            for c in db.query(models.ProductCategory).order_by(models.ProductCategory.name).all()]


@router.post("/categories")
def create_category(body: CategoryIn, db: Session = Depends(get_db)):
    c = models.ProductCategory(name=body.name.strip())
    db.add(c); db.commit(); db.refresh(c)
    return {"id": c.id}


# ══ Party base (suppliers/customers) ══════════════════════════════════════════
class PartyIn(BaseModel):
    name: str
    phone: Optional[str] = ""
    email: Optional[str] = ""
    address: Optional[str] = ""
    gst_number: Optional[str] = ""


def _party_dict(p):
    return {"id": p.id, "name": p.name, "phone": p.phone, "email": p.email,
            "address": p.address, "gst_number": p.gst_number}


def _make_party_routes(path, Model):
    @router.get(f"/{path}", name=f"list_{path}")
    def _list(db: Session = Depends(get_db)):
        rows = db.query(Model).filter(Model.is_active == 1).order_by(Model.name).all()
        return [_party_dict(p) for p in rows]

    @router.post(f"/{path}", name=f"create_{path}")
    def _create(body: PartyIn, db: Session = Depends(get_db)):
        p = Model(name=body.name.strip(), phone=body.phone, email=body.email,
                  address=body.address, gst_number=(body.gst_number or "").upper())
        db.add(p); db.commit(); db.refresh(p)
        return {"id": p.id}

    @router.put(f"/{path}/{{pid}}", name=f"update_{path}")
    def _update(pid: int, body: PartyIn, db: Session = Depends(get_db)):
        p = db.query(Model).get(pid)
        if not p: raise HTTPException(404, "Not found")
        p.name, p.phone, p.email = body.name.strip(), body.phone, body.email
        p.address, p.gst_number = body.address, (body.gst_number or "").upper()
        db.commit(); return {"ok": True}

    @router.delete(f"/{path}/{{pid}}", name=f"delete_{path}")
    def _delete(pid: int, db: Session = Depends(get_db)):
        p = db.query(Model).get(pid)
        if p:
            p.is_active = 0          # soft delete (referenced by bills/orders)
            db.commit()
        return {"ok": True}


_make_party_routes("suppliers", models.Supplier)
_make_party_routes("customers", models.Customer)


# Which spreadsheet headings map to which customer field.
IMPORT_HEADERS = {
    "name": ["name", "customer", "party", "shop", "company"],
    "phone": ["phone", "mobile", "contact", "mob", "number"],
    "email": ["email", "e-mail", "mail"],
    "address": ["address", "add.", "city", "location", "area"],
    "gst_number": ["gst", "gstin"],
}


@router.post("/customers/import")
async def import_customers(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Bulk-add customers from an Excel (.xlsx) or CSV file. The first row can
    be headings (Name / Phone / Email / Address / GST in any order); without
    headings the columns are taken as: Name, Phone, Email, Address, GST."""
    fname = (file.filename or "").lower()
    content = await file.read()

    rows = []
    if fname.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception:
            raise HTTPException(400, "Could not read this Excel file — save it as .xlsx and try again")
        for r in wb.active.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c).strip() for c in r])
    elif fname.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        rows = [[(c or "").strip() for c in r] for r in csv.reader(io.StringIO(text))]
    else:
        raise HTTPException(400, "Upload an Excel (.xlsx) or CSV file")

    rows = [r for r in rows if any(c for c in r)]
    if not rows:
        raise HTTPException(400, "The file is empty")

    # Detect a heading row and map columns; otherwise use positional order.
    col_map = {}
    for idx, cell in enumerate([c.lower() for c in rows[0]]):
        if not cell:
            continue
        for field, keys in IMPORT_HEADERS.items():
            if field not in col_map.values() and any(k in cell for k in keys):
                col_map[idx] = field
                break
    if col_map:
        data = rows[1:]
    else:
        col_map = {0: "name", 1: "phone", 2: "email", 3: "address", 4: "gst_number"}
        data = rows

    existing = {c.name.strip().lower() for c in
                db.query(models.Customer).filter(models.Customer.is_active == 1).all()}
    added, dup, invalid, added_names = 0, 0, 0, []
    for r in data:
        rec = {f: "" for f in IMPORT_HEADERS}
        for idx, field in col_map.items():
            if idx < len(r):
                rec[field] = r[idx]
        name = rec["name"].strip()
        if not name:
            invalid += 1
            continue
        if name.lower() in existing:
            dup += 1
            continue
        db.add(models.Customer(name=name, phone=rec["phone"], email=rec["email"],
                               address=rec["address"], gst_number=rec["gst_number"].upper()))
        existing.add(name.lower())
        added += 1
        if len(added_names) < 100:
            added_names.append(name)
    db.commit()
    return {"added": added, "skipped_duplicate": dup, "skipped_no_name": invalid,
            "names": added_names}


@router.get("/customers/{pid}/summary")
def customer_summary(pid: int, db: Session = Depends(get_db)):
    """Everything sold to this customer: each order form (by its reference
    number), the designs on it, and delivery progress + delivery references."""
    cust = db.query(models.Customer).get(pid)
    if not cust:
        raise HTTPException(404, "Not found")

    bills = (db.query(models.SalesBill).filter_by(customer_id=pid)
             .order_by(models.SalesBill.id.desc()).all())
    out = []
    for b in bills:
        items = db.query(models.SalesBillItem).filter_by(bill_id=b.id).all()
        # Delivery progress lives on the mirrored order.
        delivered_by_design = {}
        delivery_refs = []
        total_delivered = 0.0
        if b.order_id:
            for it in db.query(models.OrderItem).filter_by(order_id=b.order_id).all():
                delivered_by_design[(it.design_no or "").lower()] = it.delivered_qty or 0
                total_delivered += it.delivered_qty or 0
            delivery_refs = [
                {"reference_no": d.reference_no or "", "date": d.delivery_date,
                 "design_no": d.design_no or "", "pieces": d.pieces or 0}
                for d in db.query(models.OrderDelivery).filter_by(order_id=b.order_id)
                    .order_by(models.OrderDelivery.id.desc()).all()]
        total_qty = b.total_qty or 0
        out.append({
            "bill_id": b.id, "bill_number": b.bill_number,
            "reference_no": b.reference_no or "", "bill_date": b.bill_date,
            "total_qty": total_qty, "total_amount": b.total_amount or 0,
            "delivered_qty": total_delivered,
            "completed": total_qty > 0 and total_delivered >= total_qty,
            "items": [{"design_no": it.design_no or "", "qty": it.row_qty or 0,
                       "delivered": delivered_by_design.get((it.design_no or "").lower(), 0),
                       "amount": it.amount or 0} for it in items],
            "delivery_refs": delivery_refs,
        })
    return {"customer": cust.name, "phone": cust.phone or "", "bills": out}


# ══ Raw material types ════════════════════════════════════════════════════════
class MaterialTypeIn(BaseModel):
    name: str
    unit_id: Optional[int] = None
    low_stock_threshold: float = 0
    description: Optional[str] = ""


@router.get("/material-types")
def list_material_types(db: Session = Depends(get_db)):
    out = []
    for m in db.query(models.RawMaterialType).order_by(models.RawMaterialType.name).all():
        unit = db.query(models.Unit).get(m.unit_id) if m.unit_id else None
        out.append({"id": m.id, "name": m.name, "unit_id": m.unit_id,
                    "unit": unit.name if unit else "",
                    "low_stock_threshold": m.low_stock_threshold,
                    "description": m.description})
    return out


@router.post("/material-types")
def create_material_type(body: MaterialTypeIn, db: Session = Depends(get_db)):
    m = models.RawMaterialType(name=body.name.strip(), unit_id=body.unit_id,
                               low_stock_threshold=body.low_stock_threshold,
                               description=body.description)
    db.add(m); db.commit(); db.refresh(m)
    return {"id": m.id}


@router.put("/material-types/{mid}")
def update_material_type(mid: int, body: MaterialTypeIn, db: Session = Depends(get_db)):
    m = db.query(models.RawMaterialType).get(mid)
    if not m: raise HTTPException(404, "Not found")
    m.name, m.unit_id = body.name.strip(), body.unit_id
    m.low_stock_threshold, m.description = body.low_stock_threshold, body.description
    db.commit(); return {"ok": True}


@router.delete("/material-types/{mid}")
def delete_material_type(mid: int, db: Session = Depends(get_db)):
    m = db.query(models.RawMaterialType).get(mid)
    if m:
        try:
            db.delete(m); db.commit()
        except Exception:
            db.rollback(); raise HTTPException(409, "Material type is in use")
    return {"ok": True}


# ══ Products ══════════════════════════════════════════════════════════════════
class ProductIn(BaseModel):
    name: str
    category_id: Optional[int] = None
    unit_id: Optional[int] = None
    sale_rate: float = 0          # base rate (fallback for sizes with no rate)
    rate_s: float = 0
    rate_m: float = 0
    rate_l: float = 0
    rate_xl: float = 0
    rate_xxl: float = 0
    rate_mxxl: float = 0
    rate_xxxl: float = 0
    rate_xxxxl: float = 0
    description: Optional[str] = ""


_RATE_COLS = ["rate_s", "rate_m", "rate_l", "rate_xl", "rate_xxl",
              "rate_xxxl", "rate_xxxxl", "rate_mxxl"]


def _size_rates(p):
    return {c: getattr(p, c) or 0 for c in _RATE_COLS}


@router.get("/products")
def list_products(db: Session = Depends(get_db)):
    out = []
    for p in db.query(models.Product).filter(models.Product.is_active == 1)\
            .order_by(models.Product.name).all():
        cat = db.query(models.ProductCategory).get(p.category_id) if p.category_id else None
        unit = db.query(models.Unit).get(p.unit_id) if p.unit_id else None
        fg = db.query(models.FinishedGoodsStock).filter_by(product_id=p.id).first()
        out.append({"id": p.id, "name": p.name, "category_id": p.category_id,
                    "category": cat.name if cat else "", "unit_id": p.unit_id,
                    "unit": unit.name if unit else "", "sale_rate": p.sale_rate,
                    "description": p.description,
                    "pending_qty": p.pending_qty or 0,
                    "finished_qty": fg.quantity if fg else 0,
                    **_size_rates(p)})
    return out


def _apply_product(p, body: "ProductIn"):
    p.name, p.category_id, p.unit_id = body.name.strip(), body.category_id, body.unit_id
    p.sale_rate, p.description = body.sale_rate, body.description
    for c in _RATE_COLS:
        setattr(p, c, getattr(body, c))


@router.post("/products")
def create_product(body: ProductIn, db: Session = Depends(get_db)):
    p = models.Product()
    _apply_product(p, body)
    db.add(p); db.commit(); db.refresh(p)
    return {"id": p.id}


@router.put("/products/{pid}")
def update_product(pid: int, body: ProductIn, db: Session = Depends(get_db)):
    p = db.query(models.Product).get(pid)
    if not p: raise HTTPException(404, "Not found")
    _apply_product(p, body)
    db.commit(); return {"ok": True}


@router.delete("/products/{pid}")
def delete_product(pid: int, db: Session = Depends(get_db)):
    p = db.query(models.Product).get(pid)
    if p:
        p.is_active = 0          # soft delete (referenced by bills)
        db.commit()
    return {"ok": True}


class SetRateIn(BaseModel):
    rate: float


@router.post("/products/{pid}/set-rate")
def set_rate(pid: int, body: SetRateIn, db: Session = Depends(get_db)):
    """Price a job-work return: sets the product's sale rate and moves any
    pending (awaiting-rate) quantity into Finished Goods stock."""
    p = db.query(models.Product).get(pid)
    if not p:
        raise HTTPException(404, "Not found")
    if body.rate <= 0:
        raise HTTPException(400, "Enter a valid rate")

    p.sale_rate = body.rate
    pending = p.pending_qty or 0
    if pending > 0:
        services.adjust_finished_stock(db, p.id, pending)
        db.add(models.FinishedGoodsTransaction(
            product_id=p.id, transaction_type="jobwork_priced", quantity=pending,
            reference_type="set_rate"))
        p.pending_qty = 0

    db.commit()
    return {"ok": True, "moved_to_finished": pending}
